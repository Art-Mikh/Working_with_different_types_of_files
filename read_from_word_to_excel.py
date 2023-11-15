"""
File containing the main logic for
writing data from a Word file to Excel
"""
from docx import Document  # python-docx
from openpyxl import load_workbook
from dataclasses import dataclass
from file_handler import FileHandler as file_handler


@dataclass
class ExcelParametrs:
    """
    Class for storing data about an excel file
    Class entry point: main()
    atrs:
        file_path: the path to the file
        work_book: stores excel book
        working_page: stores excel sheet
    """
    file_path: str
    work_book: load_workbook
    working_page: load_workbook


class WordToExcel:
    """
    The class reads data line
    by line from a Word document
    and writes it to an Excel file
    in the same sequence
    """
    def main(self) -> None:
        """
        Starting point of the program for
        extracting data from Word to Excel
        """
        self.get_data_from_word_file()

    def get_data_from_word_file(self) -> None:
        """
        The method reads data from a Word document
        """
        file_types = (("Word file", "*.docx"),)
        file_path = file_handler.get_file_path(file_types)
        if not file_path:
            self.display_type_error_message_Wd()
            return None
        try:
            word_doc = Document(file_path)
        except Exception as error_text:
            self.display_any_error_message_Wd(error_text)

        self.processing_received_data(word_doc)

    def processing_received_data(self, word_doc: Document) -> None:
        """
        A method to write the received data
        line by line from Word to Excel
        Args:
            word_doc (docx.Document):
            data read from a Word document
        """
        self.checking_word_parameter_type(word_doc)

        excel_parametrs = self.open_excel_file()
        try:
            if excel_parametrs:
                file_path = excel_parametrs.file_path
                work_book = excel_parametrs.work_book
                working_page = excel_parametrs.working_page

                for line in word_doc.paragraphs:
                    working_page.append([str(line.text)])
                work_book.save(file_path)
                work_book.close()
                self.display_success_message()
        except TypeError as error_text:
            work_book.close()
            self.display_any_error_message_Ex(error_text)

    @staticmethod
    def checking_word_parameter_type(word_doc: Document) -> None:
        if str(type(word_doc)) != "<class 'docx.document.Document'>":
            raise TypeError("Тип передаваемого параметра не 'Document'")

    def open_excel_file(self) -> ExcelParametrs:
        """
        The method opens an Excel file and
        creates class objects for Excel manipulation
        Returns:
            ExcelParametrs:
            class object for manipulating Excel file
        """
        file_types = (("Excel file", "*.xlsx"),)
        file_path = file_handler.get_file_path(file_types)
        return self.get_excel_parameters(file_path)

    def get_excel_parameters(self, file_path: str) -> ExcelParametrs:
        """
        Method for creating and returning
        an object with Excel file parameters
        Args:
            file_path (str): the path to the file
        Raises:
            TypeError: error handler when working with a file
        Returns:
            ExcelParametrs: excel file settings
        """
        self.check_file_path_type(file_path)
        try:
            if not file_path:
                raise TypeError
            work_book = load_workbook(file_path)
            working_page = work_book["Лист1"]
            excel_data = ExcelParametrs(
                file_path=file_path,
                work_book=work_book,
                working_page=working_page
            )
            return excel_data
        except TypeError:
            self.display_type_error_message_Ex()
        except Exception as error_text:
            work_book.close()
            self.display_any_error_message_Ex(str(error_text.__str__))
        return None

    @staticmethod
    def check_file_path_type(file_path: str) -> None:
        """
        Method to check whether
        a file path is a string or not
        Args:
            file_path (str): full path to the file
        Raises:
            TypeError: called when thedata type is incorrect
        """
        if not isinstance(file_path, str):
            raise TypeError("Должна передаваться строка, содержащая путь к файлу")

    @classmethod
    def display_type_error_message_Wd(cls) -> None:
        """
        Method for displaying an error
        if a Word document was not selected
        """
        file_handler.display_file_message(
            "Ошибка при выборе файла Word!",
            "Ошибка! Word-документ не был выбран!"
        )

    @classmethod
    def display_any_error_message_Wd(cls, error_text: str = '') -> None:
        """
        Method for displaying an error when reading a Word document
        Args:
            error_text (str, optional): error text
        """
        cls.check_error_message(error_text)

        file_handler.display_file_message(
            "Ошибка при чтении Word файла",
            f"Ошибка! Word-документ прочитан с ошибкой!\n {error_text}"
        )

    @classmethod
    def display_type_error_message_Ex(cls) -> None:
        """
        Method that displays an error when an Excel file is not selected
        Args:
            error_text (str, optional): error text
        """
        file_handler.display_file_message(
            "Файл Excel не был выбран",
            "Ошибка! Файл Excel не был выбран!"
        )

    @classmethod
    def display_any_error_message_Ex(cls, error_text: str = '') -> None:
        """
        Method for displaying an error when processing an Excel file
        Args:
            error_text (str, optional):  error text
        """
        cls.check_error_message(error_text)

        file_handler.display_file_message(
            "Ошибка при обработке файла Excel",
            f"Ошибка! Excel не обработан!\n {error_text}"
        )

    @staticmethod
    def check_error_message(error_text: str = '') -> None:
        if not isinstance(error_text, str):
            raise TypeError("Должна передаваться строка, содержащая текст ошибки")

    @classmethod
    def display_success_message(cls) -> None:
        """
        Method that displays a message
        about successful data writing
        """
        file_handler.display_file_message(
            "Запись прошла успешно!",
            "Запись прошла успешно!"
        )
